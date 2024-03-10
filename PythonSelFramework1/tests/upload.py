from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

def update_excel_data(filepath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    Data = {}

    for i in range(1,sheet.max_column+ 1):
        if sheet.cell(row=1,column=i).value == colName:
            Data["col"] = i

    for i in range(1,sheet.max_row+ 1):
        for j in range(1,sheet.max_column+ 1):
            if sheet.cell(row=i,column=j).value == searchTerm:
                Data["row"] = i
    sheet.cell(row=Data["row"], column=Data["col"]).value = new_value
    book.save(filepath)




file_path = "C:\\Users\\rphilip\\Downloads\\download.xlsx"
fruit_name = "Apple"
newValue = "999"
service_obj = Service("/Users/rphilip/Documents/BrowserDriver/msedgedriver.exe")
driver = webdriver.Edge(service=service_obj)
driver.implicitly_wait(4)
driver.maximize_window()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.find_element(By.ID,"downloadButton").click()
driver.implicitly_wait(4)
#edit the excel with updated value
update_excel_data(file_path,"Apple","price",newValue)

file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver,5)
toast_locater = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toast_locater))
print(driver.find_element(*toast_locater).text)

priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
assert actual_price == newValue

