import openpyxl

book = openpyxl.load_workbook("C:/Users/rphilip/Downloads/download.xlsx")

sheet = book.active


Data = []
new_value = 945

for i in range(2,sheet.max_row+1):
    #if sheet.cell(row=i,column=1).value == "Testcase1":
    Dict = {}
    for j in range(2,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)
        #Dict["lastname"] = "shetty"

        Dict[sheet.cell(row =1,column=j).value] = sheet.cell(row =i,column=j).value
        #print(Dict)
    Data = Data + [Dict]
print(Data)
