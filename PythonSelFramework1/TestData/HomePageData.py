import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"Rahul","lastname":"shetty","gender":"Male"},{"firstname":"Anshika","lastname":"shetty","gender":"Female"}]

    @staticmethod
    def getTestData():

        Data =[]
        book = openpyxl.load_workbook("C:\\Users\\rphilip\\PycharmProjects\\Excel\\PythonDemo.xlsx")

        sheet = book.active
        for i in range(2, sheet.max_row + 1):
            #if sheet.cell(row=i, column=1).value == test_case_name:
            Dict = {}
            for j in range(2, sheet.max_column + 1):
                # Dict["lastname"] = "shetty"
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            Data = Data + [Dict]
        return Data
