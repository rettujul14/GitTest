import openpyxl

book = openpyxl.load_workbook("C:\\Users\\rphilip\\PycharmProjects\\Excel\\PythonDemo.xlsx")

#sheet = book.active
sheet = book["Test"]
#print(book.sheetnames)


Data = []
t = 0

cell = sheet.cell(row=1,column=2)

print(cell.value)

sheet.cell(row=2,column=2).value = "Rahul"

print(sheet.cell(row=2,column=2).value )

print(sheet.max_row)

print(sheet.max_column)

print(sheet['A3'].value)

for i in range(2,sheet.max_row+1):
    #if sheet.cell(row=i,column=1).value == "Testcase1":
    Dict = {}
    for j in range(2,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)
        #Dict["lastname"] = "shetty"

        Dict[sheet.cell(row =1,column=j).value] = sheet.cell(row =i,column=j).value
        #print(Dict)
    Data = Data + [Dict]
print(Data)
